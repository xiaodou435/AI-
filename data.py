from zhushou import base
from openpyxl import load_workbook
# dianzhan = base("军营村电站2024年4月24日收益")
# dianzhan = base("介绍一下有负荷测算")

file = "D:\AI助手\平台问题.xlsx"
# load file 
wb = load_workbook(file)

sheet = wb.active
#read A1 data
# cell_value = sheet['A1'].value
# print(f"Value in cell A1:{cell_value}")
# read A column data

column_a = sheet['A']
for cell in column_a:
    print(cell.value)
    base(cell.value)

#read all data
# for row in sheet.iter_rows(values_only=True):
#     print(row)
wb.close()
