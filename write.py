from openpyxl import load_workbook, Workbook


def w_d(data):
    # 指定要写入的 Excel 文件名
    output_file = 'output.xlsx'

    # 检查文件是否存在
    try:
        wb = load_workbook(output_file)
        print(f"Loading existing workbook {output_file}")
    except FileNotFoundError:
        wb = Workbook()
        print(f"Creating a new workbook {output_file}")

    # 选择第一个工作表，如果文件是新创建的，默认第一个工作表已经存在
    sheet = wb.active

    # 清空工作表，重新写入数据
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None

    # 将数据写入 Excel 文件
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    # 保存工作簿
    wb.save(output_file)
    print(f"Data has been written to {output_file}")
    wb.close()
data = [
    ["Header1", "Header2", "Header3"],
    ["Row1Value1", "Row1Value2", "Row1Value3"],
    ["Row2Value1", "Row2Value2", "Row2Value3"]
]
# w_d(data)